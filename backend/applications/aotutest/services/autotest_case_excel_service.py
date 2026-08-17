# -*- coding: utf-8 -*-
"""
@Author  : yangkai
@Email   : 807440781@qq.com
@Project : Krun
@Module  : autotest_case_excel_service.py
@DateTime: 2026/8/1
"""
from __future__ import annotations

import io
import json
import os
import re
from collections import Counter
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from tortoise.transactions import in_transaction

from backend.applications.aotutest.models.autotest_case_model import AutoTestCaseModel
from backend.applications.aotutest.models.autotest_project_model import AutoTestProjectModel
from backend.applications.aotutest.models.autotest_step_model import AutoTestStepModel
from backend.configure import LOGGER, PROJECT_CONFIG
from backend.enums import (
    AutoTestAssertionOperation,
    AutoTestCaseAttr,
    AutoTestCaseType,
    AutoTestReqArgsType,
    AutoTestStepType,
    HTTPMethod,
)
from backend.services import get_current_username

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

_HTTP, _TCP = "HTTP", "TCP"
_MARKER_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ROW_HEIGHT = 40
_COL_WIDTH_MIN, _COL_WIDTH_MAX = 8, 60

_SCRIPT_COLUMNS: Tuple[str, ...] = (
    "接口名称", "所属应用", "协议类型", "接口描述",
    "请求方式", "配置名称", "请求路径", "请求体类型", "请求体", "请求头",
    "变量", "提取", "断言", "所属人员",
)
_DATA_START_ROW = 3
_SCOPE_ALL, _SCOPE_SOME = "整个返回数据", "提取部分"
_EXTRACT_SOURCES = frozenset({
    "Request Form-Data", "Request Text", "Request Json", "Request XML", "Request Headers",
    "Response Text", "Response Json", "Response XML", "Response Headers", "Response Cookie",
})
_ASSERT_SOURCES = frozenset({
    "Response Text", "Response Json", "Response XML", "Response Headers", "Response Cookie", "变量池",
})
_ASSERT_OPS = frozenset(e.value for e in AutoTestAssertionOperation)
_HTTP_METHODS = frozenset(e.value for e in HTTPMethod)
_ARGS_TYPES = frozenset(e.value for e in AutoTestReqArgsType)
_TCP_ARGS = frozenset({
    AutoTestReqArgsType.XML.value, AutoTestReqArgsType.JSON.value, AutoTestReqArgsType.RAW.value,
})
# 表单类请求体：args_type → 步骤字段名（导出矩阵/脚本序列化/往返检测共用）
_FORM_ATTR = {
    AutoTestReqArgsType.FORM_DATA.value: "request_form_data",
    AutoTestReqArgsType.X_WWW_FORM_URLENCODED.value: "request_form_urlencoded",
    AutoTestReqArgsType.PARAMS.value: "request_params",
}
_SCRIPT_TEMPLATE = os.path.join(PROJECT_CONFIG.OUTPUT_DIR, "template", "公共接口导入导出模板.xlsx")


# ---------------------------------------------------------------------------
# 公共工具
# ---------------------------------------------------------------------------

def _display_text_width(text: str) -> int:
    """
    估算单元格显示宽度：ASCII计1，宽字符计2。
    """
    return sum(2 if ord(ch) > 127 else 1 for ch in text)


def _auto_size_sheet_columns(sheet) -> None:
    """
    根据单元格内容自适应列宽，多行文本取最长行。
    """
    max_column = sheet.max_column or 0
    max_row = sheet.max_row or 0
    for col_idx in range(1, max_column + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            for line in str(value).splitlines() or [""]:
                max_len = max(max_len, _display_text_width(line))
        width = min(_COL_WIDTH_MAX, max(_COL_WIDTH_MIN, max_len + 2))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _style_sheet_cells(sheet, *, start_row: int = 1, row_height: Optional[float] = _ROW_HEIGHT) -> None:
    """
    数据区水平/垂直居中；可选统一行高。
    """
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0
    if max_row < start_row or max_column < 1:
        return
    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, max_col=max_column):
        if row_height is not None:
            sheet.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.alignment = _CENTER_ALIGN


def _file_name(username: Optional[str], label: str) -> str:
    safe = re.sub(r'[\\/:*?"<>|\s]', "_", str(username or "").strip())
    prefix = f"{safe}_{label}" if safe else label
    return f"{prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"


def _get(item: Any, name: str, default: Any = None) -> Any:
    """
    兼容dict/schema对象字段读取。
    """
    return item.get(name, default) if isinstance(item, dict) else getattr(item, name, default)


def _enum_val(raw: Any) -> str:
    """CharEnum / 字符串 → 枚举值文本。"""
    if raw is None:
        return ""
    return getattr(raw, "value", raw) or ""


def _collect_own_steps(steps: Optional[List[Any]]) -> List[Any]:
    collected: List[Any] = []
    for step in steps or []:
        collected.append(step)
        collected.extend(_collect_own_steps(getattr(step, "children", None)))
    return collected


async def _load_public_api_cases(case_ids: List[int], services: Any) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    两通道共用：公共接口 + 唯一步骤HTTP/TCP + 无数据源。
    """
    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    project_names: Dict[Any, str] = {}
    for case_id in case_ids:
        case = await services.case_curd.get_by_id(case_id)
        case_name = (getattr(case, "case_name", None) or str(case_id)) if case else str(case_id)
        if not case:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "用例不存在"})
            continue
        if getattr(case, "case_type", None) != AutoTestCaseType.PUBLIC_API:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "非公共接口用例"})
            continue
        load = await services.step_curd.get_by_case_id(case_id=case_id)
        own_steps = _collect_own_steps(getattr(load, "root_steps", None))
        if len(own_steps) != 1:
            invalid.append({
                "case_id": case_id, "case_name": case_name,
                "reason": f"用例步骤数为{len(own_steps)}，需且仅需1步",
            })
            continue
        step = own_steps[0]
        step_type = getattr(step, "step_type", None)
        if step_type not in (AutoTestStepType.HTTP, AutoTestStepType.TCP):
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "步骤非HTTP/TCP请求步骤"})
            continue
        if getattr(step, "data_source_id", None):
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "步骤存在数据源绑定"})
            continue
        project_id = getattr(case, "case_project", None)
        if project_id not in project_names:
            project = await services.project_curd.get_by_id(project_id) if project_id else None
            project_names[project_id] = (getattr(project, "project_name", None) or "") if project else ""
        valid.append({
            "case_id": case_id,
            "case": case,
            "case_name": case_name,
            "step": step,
            "is_http": step_type == AutoTestStepType.HTTP,
            "project_name": project_names[project_id],
        })
    return valid, invalid


# ---------------------------------------------------------------------------
# 通道 A：导出数据（HEAD/BODY JSONPath 矩阵）
# ---------------------------------------------------------------------------

def build_export_file_name(username: Optional[str]) -> str:
    return _file_name(username, "接口报文数据")


def _kv_to_pairs(kv_list: Optional[List[Any]]) -> List[Tuple[str, Any]]:
    pairs: List[Tuple[str, Any]] = []
    for item in kv_list or []:
        key = _get(item, "key")
        if key is None or not str(key).strip():
            continue
        pairs.append((f"$.{str(key).strip()}", _get(item, "value")))
    return pairs


def _flatten_jsonpath(data: Any, prefix: str = "$") -> List[Tuple[str, Any]]:
    pairs: List[Tuple[str, Any]] = []
    if isinstance(data, dict):
        for key, value in data.items():
            path = f"{prefix}.{key}"
            if isinstance(value, (dict, list)) and value:
                pairs.extend(_flatten_jsonpath(value, path))
            else:
                pairs.append((path, "" if isinstance(value, (dict, list)) else value))
    elif isinstance(data, list):
        for index, value in enumerate(data):
            path = f"{prefix}[{index}]"
            if isinstance(value, (dict, list)) and value:
                pairs.extend(_flatten_jsonpath(value, path))
            else:
                pairs.append((path, "" if isinstance(value, (dict, list)) else value))
    return pairs


def _xml_local_name(tag: str) -> str:
    """去掉 Clark 命名空间，仅保留本地标签名。"""
    if tag and "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return tag or ""


def _flatten_xml_to_jsonpath_pairs(xml_text: str) -> List[Tuple[str, Any]]:
    """
    按XML文档序将叶子字段展平为JSONPath列（不经xmltodict，保留空标签与声明序）。

    路径约定与JSON展平对齐：$.Root.Child；同名兄弟仅在重复时加0-based下标；
    属性为$.Path.@attr。非法XML回退$.raw。
    """
    text = str(xml_text or "").strip()
    if not text:
        return []
    try:
        root = ElementTree.fromstring(text)
    except ElementTree.ParseError:
        return [("$.raw", xml_text)]

    pairs: List[Tuple[str, Any]] = []

    def walk(elem: ElementTree.Element, path: str) -> None:
        for attr_name, attr_value in elem.attrib.items():
            local_attr = _xml_local_name(attr_name)
            pairs.append((f"{path}.@{local_attr}", "" if attr_value is None else attr_value))

        children = list(elem)
        if not children:
            pairs.append((path, "" if elem.text is None else elem.text))
            return

        name_total = Counter(_xml_local_name(child.tag) for child in children)
        name_seen: Counter = Counter()
        for child in children:
            local = _xml_local_name(child.tag)
            name_seen[local] += 1
            if name_total[local] > 1:
                child_path = f"{path}.{local}[{name_seen[local] - 1}]"
            else:
                child_path = f"{path}.{local}"
            walk(child, child_path)

    root_name = _xml_local_name(root.tag) or "root"
    walk(root, f"$.{root_name}")
    return pairs


def _body_to_pairs(step: Any) -> List[Tuple[str, Any]]:
    args = getattr(step, "request_args_type", None)
    if args == AutoTestReqArgsType.JSON:
        body = getattr(step, "request_body", None)
        return _flatten_jsonpath(body) if isinstance(body, dict) else []
    if args == AutoTestReqArgsType.XML:
        text = getattr(step, "request_text", None) or ""
        return _flatten_xml_to_jsonpath_pairs(text) if text else []
    if args == AutoTestReqArgsType.RAW:
        text = getattr(step, "request_text", None) or ""
        return [("$.raw", text)] if text else []
    attr = _FORM_ATTR.get(_enum_val(args))
    return _kv_to_pairs(getattr(step, attr, None)) if attr else []


async def prepare_export_cases(case_ids: List[int], services: Any) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    loaded, invalid = await _load_public_api_cases(case_ids, services)
    valid = [{
        "case_name": item["case_name"],
        "protocol": _HTTP if item["is_http"] else _TCP,
        "project_name": item["project_name"],
        "case_desc": getattr(item["case"], "case_desc", None) or "",
        "created_user": getattr(item["case"], "created_user", None) or "",
        "head_pairs": _kv_to_pairs(getattr(item["step"], "request_header", None)),
        "body_pairs": _body_to_pairs(item["step"]),
    } for item in loaded]
    LOGGER.info(f"导出用例准备完成: 有效{len(valid)}个, 不合规{len(invalid)}个")
    return valid, invalid


def build_export_workbook(cases_data: List[Dict[str, Any]]) -> Workbook:
    """
    根据一用例一sheet构建导出工作簿，多于1个时前置目录sheet。
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set = set()
    sheet_titles: List[Tuple[str, Dict[str, Any]]] = []

    for case_data in cases_data:
        raw = re.sub(r"[:\\/?*\[\]]", "_", str(case_data.get("case_name") or "").strip()) or "用例"
        title, base, idx = raw[:31], raw[:31], 1
        while title in used_names:
            suffix = f"_{idx}"
            title = base[:31 - len(suffix)] + suffix
            idx += 1
        used_names.add(title)

        sheet = workbook.create_sheet(title=title)
        header: List[Any] = []
        values: List[Any] = []
        for label, field in (("HEAD", "head_pairs"), ("BODY", "body_pairs")):
            header.append(label)
            values.append("")
            for path, value in case_data.get(field) or []:
                header.append(path)
                values.append("" if value is None else (value if isinstance(value, (str, int, float, bool)) else str(value)))
        sheet.append(header)
        sheet.append(values)
        sheet_titles.append((title, case_data))

    if len(sheet_titles) > 1:
        directory = workbook.create_sheet(title="目录", index=0)
        directory.append(["序号", "接口名称", "所属应用", "接口描述", "所属人", "协议类型"])
        for index, (title, case_data) in enumerate(sheet_titles, start=1):
            directory.append([
                index, case_data.get("case_name"), case_data.get("project_name"),
                case_data.get("case_desc"), case_data.get("created_user"), case_data.get("protocol"),
            ])
            cell = directory.cell(row=index + 1, column=2)
            cell.hyperlink = f"#'{title}'!A1"
            cell.style = "Hyperlink"

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ("HEAD", "BODY"):
                    cell.fill = _MARKER_FILL
        _style_sheet_cells(sheet, start_row=1, row_height=_ROW_HEIGHT)
        _auto_size_sheet_columns(sheet)
    return workbook


# ---------------------------------------------------------------------------
# 通道 B：导出/导入脚本（15 列模板）
# ---------------------------------------------------------------------------

def build_script_file_name(username: Optional[str]) -> str:
    return _file_name(username, "接口详情数据")


def _kv_to_lines(
        kv_list: Optional[List[Any]], *, column: str = "", problems: Optional[List[str]] = None
) -> str:
    """
    将[{key,value,desc}]转为多行key:value[:desc];文本，可选做往返安全检测。
    """
    lines: List[str] = []
    for item in kv_list or []:
        key = str(_get(item, "key") or "").strip()
        if not key:
            continue
        raw_value = _get(item, "value")
        value = "" if raw_value is None else str(raw_value)
        desc = str(_get(item, "desc") or "").strip()
        if problems is not None:
            if ":" in key or "\n" in key:
                problems.append(f"「{column}」键({key})含冒号或换行, 模板格式无法安全往返")
            if ":" in value or "\n" in value:
                problems.append(f"「{column}」键({key})的值含冒号或换行, 导出后再导入会被截断错位")
            if "\n" in desc:
                problems.append(f"「{column}」键({key})的描述含换行, 模板格式无法安全往返")
        seg = f"{key}:{value}" + (f":{desc}" if desc else "")
        lines.append(seg + ";")
    return "\n".join(lines)


def _extract_to_lines(extract_list: Optional[List[Any]]) -> str:
    lines: List[str] = []
    for item in extract_list or []:
        name = str(_get(item, "name") or "").strip()
        source = str(_get(item, "source") or "").strip()
        if not name or not source:
            continue
        if _get(item, "scope") == "ALL":
            lines.append(f"{name}:{source}:{_SCOPE_ALL};")
            continue
        expr = str(_get(item, "expr") or "").strip()
        seg = f"{name}:{source}:{_SCOPE_SOME}:{expr}"
        index = _get(item, "index")
        if index is not None:
            seg += f":{int(index)}"
        lines.append(seg + ";")
    return "\n".join(lines)


def _assert_to_lines(assert_list: Optional[List[Any]]) -> str:
    lines: List[str] = []
    for item in assert_list or []:
        name = str(_get(item, "name") or "").strip()
        source = str(_get(item, "source") or "").strip()
        expr = str(_get(item, "expr") or "").strip()
        operation = str(_get(item, "operation") or "").strip()
        if not name or not source or not expr or not operation:
            continue
        raw_except = _get(item, "except_value")
        except_value = "" if raw_except is None else str(raw_except)
        lines.append(f"{name}:{source}:{expr}:{operation}:{except_value};")
    return "\n".join(lines)


def _step_body_cell(step: Any) -> str:
    args = _enum_val(getattr(step, "request_args_type", None))
    if args == AutoTestReqArgsType.JSON.value:
        body = getattr(step, "request_body", None)
        return json.dumps(body, ensure_ascii=False) if isinstance(body, dict) else ""
    if args in (AutoTestReqArgsType.XML.value, AutoTestReqArgsType.RAW.value):
        return getattr(step, "request_text", None) or ""
    attr = _FORM_ATTR.get(args)
    return _kv_to_lines(getattr(step, attr, None)) if attr else ""


async def prepare_script_export_rows(case_ids: List[int], services: Any) -> Tuple[List[Dict[str, str]], List[Dict[str, Any]]]:
    loaded, invalid = await _load_public_api_cases(case_ids, services)
    rows: List[Dict[str, str]] = []
    for item in loaded:
        case, step = item["case"], item["step"]
        case_id, case_name, is_http = item["case_id"], item["case_name"], item["is_http"]
        args = _enum_val(getattr(step, "request_args_type", None))
        method = _enum_val(getattr(step, "request_method", None))

        problems: List[str] = []
        header_text = (
            _kv_to_lines(getattr(step, "request_header", None), column="请求头", problems=problems)
            if is_http else ""
        )
        variables_text = _kv_to_lines(
            getattr(step, "defined_variables", None), column="变量", problems=problems
        )
        form_attr = _FORM_ATTR.get(args)
        if form_attr:
            # 表单类请求体：检测与序列化合一；非表单类走 _step_body_cell
            body_text = _kv_to_lines(getattr(step, form_attr, None), column="请求体", problems=problems)
        else:
            body_text = _step_body_cell(step)
        for assert_item in getattr(step, "assert_validators", None) or []:
            if ":" in str(_get(assert_item, "expr") or ""):
                problems.append(
                    f"「断言」({_get(assert_item, 'name') or ''})的断言表达式含冒号, 导入时无法正确解析"
                )
        if problems:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "；".join(problems)})
            continue

        rows.append({
            "接口名称": case_name,
            "所属应用": item["project_name"],
            "协议类型": _HTTP if is_http else _TCP,
            "接口描述": getattr(case, "case_desc", None) or "",
            "请求方式": method if is_http else "",
            "配置名称": getattr(step, "request_config_name", None) or "",
            "请求路径": (getattr(step, "request_url", None) or "") if is_http else "",
            "请求体类型": args,
            "请求体": body_text,
            "请求头": header_text,
            "变量": variables_text,
            "提取": _extract_to_lines(getattr(step, "extract_variables", None)),
            "断言": _assert_to_lines(getattr(step, "assert_validators", None)),
            "所属人员": getattr(case, "owner_user", None) or getattr(case, "created_user", None) or "",
        })
    LOGGER.info(f"导出脚本准备完成: 有效{len(rows)}个, 不合规{len(invalid)}个")
    return rows, invalid


def build_script_workbook(rows: List[Dict[str, str]]) -> Workbook:
    if not os.path.isfile(_SCRIPT_TEMPLATE):
        raise RuntimeError(f"模板文件不存在: {_SCRIPT_TEMPLATE}")
    workbook = load_workbook(_SCRIPT_TEMPLATE)
    sheet = workbook[workbook.sheetnames[0]]
    header = [cell.value for cell in sheet[1][: len(_SCRIPT_COLUMNS)]]
    if header != list(_SCRIPT_COLUMNS):
        raise RuntimeError(f"模板表头已被改动，与预定义不一致: 模板={header}, 期望={list(_SCRIPT_COLUMNS)}")
    for row_index, row in enumerate(rows, start=_DATA_START_ROW):
        for col_index, column in enumerate(_SCRIPT_COLUMNS, start=1):
            sheet.cell(row=row_index, column=col_index, value=row.get(column) or "")
    # 数据区从第 3 行起：居中 + 统一行高；列宽根据全表内容自适应
    if rows:
        _style_sheet_cells(sheet, start_row=_DATA_START_ROW, row_height=_ROW_HEIGHT)
    _auto_size_sheet_columns(sheet)
    return workbook


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_kv(text: str, errors: List[str], column: str) -> Optional[List[Dict[str, str]]]:
    items: List[Dict[str, str]] = []
    for raw in text.splitlines():
        seg = raw.strip()
        if not seg:
            continue
        if seg.endswith(";"):
            seg = seg[:-1]
        parts = seg.split(":")
        key = parts[0].strip()
        if not key:
            errors.append(f"「{column}」存在缺少key的行: {raw.strip()!r}")
            continue
        items.append({
            "key": key,
            "value": parts[1] if len(parts) >= 2 else "",
            "desc": ":".join(parts[2:]).strip() if len(parts) >= 3 else "",
        })
    return items or None


def _parse_extract(text: str, errors: List[str]) -> Optional[List[Dict[str, Any]]]:
    items: List[Dict[str, Any]] = []
    for raw in text.splitlines():
        seg = raw.strip()
        if not seg:
            continue
        if seg.endswith(";"):
            seg = seg[:-1]
        parts = seg.split(":")
        if len(parts) < 3:
            errors.append(
                f"「提取」格式非法(应为 变量名:提取来源:整个返回数据 或 变量名:提取来源:提取部分:表达式): {raw.strip()!r}"
            )
            continue
        name, source, scope = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if not name:
            errors.append(f"「提取」存在缺少变量名的行: {raw.strip()!r}")
            continue
        if source not in _EXTRACT_SOURCES:
            errors.append(f"「提取」提取来源({source})非法, 合法集: {'/'.join(sorted(_EXTRACT_SOURCES))}")
            continue
        if scope == _SCOPE_ALL:
            if len(parts) > 3:
                errors.append(f"「提取」整个返回数据不应携带提取表达式: {raw.strip()!r}")
                continue
            items.append({"name": name, "source": source, "expr": "", "scope": "ALL", "index": None})
        elif scope == _SCOPE_SOME:
            tail = parts[3:]
            index_val: Optional[int] = None
            if len(tail) >= 2 and re.fullmatch(r"-?\d+", tail[-1].strip()):
                index_val = int(tail[-1].strip())
                tail = tail[:-1]
            expr = ":".join(tail).strip()
            if not expr:
                errors.append(f"「提取」部分提取缺少提取表达式: {raw.strip()!r}")
                continue
            items.append({"name": name, "source": source, "expr": expr, "scope": "SOME", "index": index_val})
        else:
            errors.append(f"「提取」提取范围({scope})须为「{_SCOPE_ALL}」或「{_SCOPE_SOME}」")
    return items or None


def _parse_assert(text: str, errors: List[str]) -> Optional[List[Dict[str, Any]]]:
    items: List[Dict[str, Any]] = []
    for raw in text.splitlines():
        seg = raw.strip()
        if not seg:
            continue
        if seg.endswith(";"):
            seg = seg[:-1]
        parts = seg.split(":")
        if len(parts) < 5:
            errors.append(f"「断言」格式非法(应为 断言名称:断言对象:断言表达式:匹配规则:预期值): {raw.strip()!r}")
            continue
        name, source, expr, operation = parts[0].strip(), parts[1].strip(), parts[2].strip(), parts[3].strip()
        if not name:
            errors.append(f"「断言」存在缺少断言名称的行: {raw.strip()!r}")
            continue
        if source not in _ASSERT_SOURCES:
            errors.append(f"「断言」断言对象({source})非法, 合法集: {'/'.join(sorted(_ASSERT_SOURCES))}")
            continue
        if not expr:
            errors.append(f"「断言」断言表达式不允许为空: {raw.strip()!r}")
            continue
        if operation not in _ASSERT_OPS:
            errors.append(f"「断言」匹配规则({operation})非法, 合法集: {'/'.join(sorted(_ASSERT_OPS))}")
            continue
        items.append({
            "name": name, "source": source, "expr": expr,
            "operation": operation, "except_value": ":".join(parts[4:]),
        })
    return items or None


def _parse_body(args_type: str, body_text: str, errors: List[str]) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "request_body": None, "request_text": None,
        "request_params": None, "request_form_data": None, "request_form_urlencoded": None,
    }
    if args_type == AutoTestReqArgsType.NONE.value:
        if body_text:
            errors.append("请求体类型为none时「请求体」不允许填写")
        return result
    if not body_text:
        errors.append(f"请求体类型为{args_type}时「请求体」不允许为空")
        return result
    if args_type == AutoTestReqArgsType.JSON.value:
        try:
            parsed = json.loads(body_text)
        except ValueError as e:
            errors.append(f"「请求体」JSON解析失败: {e}")
            return result
        if not isinstance(parsed, dict):
            errors.append("「请求体」JSON须为对象结构(以{开头)")
            return result
        result["request_body"] = parsed
    elif args_type in (AutoTestReqArgsType.XML.value, AutoTestReqArgsType.RAW.value):
        result["request_text"] = body_text
    elif args_type in _FORM_ATTR:
        result[_FORM_ATTR[args_type]] = _parse_kv(body_text, errors, "请求体")
    return result


def parse_script_workbook(content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [_cell_text(cell.value) for cell in sheet[1][: len(_SCRIPT_COLUMNS)]]
    if header != list(_SCRIPT_COLUMNS):
        return [], [{
            "row": 1,
            "reason": f"表头与模板不一致(期望前{len(_SCRIPT_COLUMNS)}列为: {'/'.join(_SCRIPT_COLUMNS)})，请使用最新模板",
        }]

    rows: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    for row_no, excel_row in enumerate(
            sheet.iter_rows(min_row=_DATA_START_ROW, max_col=len(_SCRIPT_COLUMNS), values_only=True),
            start=_DATA_START_ROW,
    ):
        cells = {col: _cell_text(val) for col, val in zip(_SCRIPT_COLUMNS, excel_row)}
        if not any(cells.values()):
            continue

        errors: List[str] = []
        case_name, project_name = cells["接口名称"], cells["所属应用"]
        protocol_raw = cells["协议类型"]
        protocol = protocol_raw.upper()
        if not case_name:
            errors.append("「接口名称」不允许为空")
        if not project_name:
            errors.append("「所属应用」不允许为空")
        if protocol not in (_HTTP, _TCP):
            errors.append(f"「协议类型」({protocol_raw})须为HTTP或TCP")
            protocol = None

        method, request_url, header_text = cells["请求方式"].upper(), cells["请求路径"], cells["请求头"]
        if protocol == _HTTP:
            if not method:
                errors.append("HTTP协议时「请求方式」不允许为空")
            elif method not in _HTTP_METHODS:
                errors.append(f"「请求方式」({method})非法, 合法集: {'/'.join(sorted(_HTTP_METHODS))}")
            if not request_url:
                errors.append("HTTP协议时「请求路径」不允许为空")
        elif protocol == _TCP:
            if method:
                errors.append("TCP协议时「请求方式」勿填")
            if request_url:
                errors.append("TCP协议时「请求路径」勿填")
            if header_text:
                errors.append("TCP协议时「请求头」禁止填写")

        args_type = cells["请求体类型"]
        if not args_type:
            errors.append("「请求体类型」不允许为空")
        elif args_type not in _ARGS_TYPES:
            errors.append(f"「请求体类型」({args_type})非法, 合法集: {'/'.join(e.value for e in AutoTestReqArgsType)}")
        elif protocol == _TCP and args_type not in _TCP_ARGS:
            errors.append(f"TCP协议时「请求体类型」({args_type})仅支持: {'/'.join(sorted(_TCP_ARGS))}")

        # 协议/请求体类型非法时跳过依赖项解析；其余列继续解析，保证单行错误一次给全
        args_ok = (
                protocol in (_HTTP, _TCP)
                and args_type in _ARGS_TYPES
                and (protocol == _HTTP or args_type in _TCP_ARGS)
        )
        body_fields = _parse_body(args_type, cells["请求体"], errors) if args_ok else {}
        request_header = _parse_kv(header_text, errors, "请求头") if protocol == _HTTP else None
        defined_variables = _parse_kv(cells["变量"], errors, "变量")
        extract_variables = _parse_extract(cells["提取"], errors)
        assert_validators = _parse_assert(cells["断言"], errors)
        if errors:
            invalid.append({"row": row_no, "reason": "；".join(errors)})
            continue
        rows.append({
            "row_no": row_no,
            "case_name": case_name,
            "project_name": project_name,
            "protocol": protocol,
            "case_desc": cells["接口描述"] or None,
            # 公共接口：步骤名称与接口名称一致（前端 Request 面板同步锁定）
            "step_name": case_name,
            "request_method": method or None,
            "request_config_name": cells["配置名称"] or None,
            "request_url": request_url or None,
            "request_args_type": args_type,
            "request_header": request_header,
            "defined_variables": defined_variables,
            "extract_variables": extract_variables,
            "assert_validators": assert_validators,
            **body_fields,
        })
    if not rows and not invalid:
        invalid.append({"row": _DATA_START_ROW, "reason": "文件无有效数据行"})
    LOGGER.info(f"导入脚本解析完成: 有效{len(rows)}行, 不合规{len(invalid)}行")
    return rows, invalid


async def import_script_rows(
        rows: List[Dict[str, Any]], services: Any
) -> Tuple[Dict[str, int], List[Dict[str, Any]]]:
    """
    按所属应用、接口名称、当前登录所属人匹配公共接口，含软删，存在则更新或恢复覆盖、不存在则新增，校验通过后单事务落库。
    """
    prepared: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    project_cache: Dict[str, Optional[int]] = {}
    owner_user = get_current_username()

    seen: Dict[Tuple[str, str], int] = {}
    duplicates: set = set()
    for row in rows:
        key = (row["project_name"], row["case_name"])
        if key in seen:
            invalid.append({
                "row": row["row_no"],
                "reason": f"文件内与第{seen[key]}行重复(同所属应用+接口名称), 无法定位唯一目标",
            })
            duplicates.add(row["row_no"])
        else:
            seen[key] = row["row_no"]

    for row in rows:
        if row["row_no"] in duplicates:
            continue
        errors: List[str] = []
        project_name = row["project_name"]
        if project_name not in project_cache:
            project = await AutoTestProjectModel.filter(project_name=project_name, state__not=1).first()
            project_cache[project_name] = project.id if project else None
        project_id = project_cache[project_name]
        if project_id is None:
            errors.append(f"所属应用({project_name})不存在")

        existing_case: Optional[AutoTestCaseModel] = None
        existing_step: Optional[AutoTestStepModel] = None
        if project_id is not None:
            if not owner_user:
                errors.append("当前登录账号为空, 无法按所属人员定位公共接口")
            else:
                matched_cases = await AutoTestCaseModel.filter(
                    case_project=project_id,
                    case_name=row["case_name"],
                    case_type=AutoTestCaseType.PUBLIC_API,
                    owner_user=owner_user,
                ).all()
                if len(matched_cases) > 1:
                    errors.append(
                        f"应用({project_name})下所属人({owner_user})存在多条同名公共接口({row['case_name']}), 无法定位"
                    )
                elif matched_cases:
                    existing_case = matched_cases[0]
                    if existing_case.state != 1:
                        root_steps = await AutoTestStepModel.filter(
                            case_id=existing_case.id, parent_step_id=None, state__not=1
                        ).all()
                    else:
                        root_steps = await AutoTestStepModel.filter(
                            case_id=existing_case.id, parent_step_id=None
                        ).all()
                    if len(root_steps) != 1:
                        errors.append(f"存量公共接口({row['case_name']})根步骤数为{len(root_steps)}, 形态异常")
                    else:
                        existing_step = root_steps[0]

        if errors:
            invalid.append({"row": row["row_no"], "reason": "；".join(errors)})
            continue
        prepared.append({
            **row,
            "project_id": project_id,
            "existing_case": existing_case,
            "existing_step": existing_step,
        })

    if invalid:
        return {}, invalid

    created = updated = 0
    async with in_transaction():
        for item in prepared:
            step_payload = {
                "step_name": item["step_name"],
                "step_desc": item["case_desc"] or "",
                "step_type": AutoTestStepType.HTTP if item["protocol"] == _HTTP else AutoTestStepType.TCP,
                "request_project_id": item["project_id"],
                "request_method": item["request_method"],
                "request_url": item["request_url"],
                "request_config_name": item["request_config_name"],
                "request_args_type": AutoTestReqArgsType(item["request_args_type"]),
                "request_body": item["request_body"],
                "request_text": item["request_text"],
                "request_header": item["request_header"],
                "request_params": item["request_params"],
                "request_form_data": item["request_form_data"],
                "request_form_urlencoded": item["request_form_urlencoded"],
                "defined_variables": item["defined_variables"],
                "extract_variables": item["extract_variables"],
                "assert_validators": item["assert_validators"],
            }
            if item["existing_case"] is None:
                new_case = await services.case_curd.create(obj_in={
                    "case_name": item["case_name"],
                    "case_project": item["project_id"],
                    "case_type": AutoTestCaseType.PUBLIC_API,
                    "case_attr": AutoTestCaseAttr.TRUE_CASE,
                    "case_desc": item["case_desc"],
                    "case_tags": None,
                    "session_variables": None,
                    "case_steps": 1,
                    "owner_user": owner_user,
                })
                await services.step_curd.create(obj_in={"case_id": new_case.id, "step_no": 1, **step_payload})
                created += 1
            else:
                existing_case = item["existing_case"]
                await services.case_curd.update(
                    id=existing_case.id,
                    obj_in={
                        "case_desc": item["case_desc"],
                        "case_tags": None,
                        "case_version": (getattr(existing_case, "case_version", None) or 1) + 1,
                        "state": 0,
                    },
                )
                await services.step_curd.update(
                    id=item["existing_step"].id,
                    obj_in={**step_payload, "state": 0},
                )
                updated += 1
    LOGGER.info(f"导入脚本落库完成: 新增{created}个, 更新{updated}个")
    return {"created_count": created, "updated_count": updated}, []
